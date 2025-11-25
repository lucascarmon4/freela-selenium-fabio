import pandas as pd
from fpdf import FPDF
import os

def salvar_excel(tarefas, horas_totais, horas_ajustadas, valor_bruto, valor_imposto, valor_liquido, pasta_projeto, valor_hora, fator, imposto_pct):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimativa do Projeto"
    
    # Estilos
    header_font = Font(bold=True, size=12)
    param_font = Font(bold=True, color="FFFFFF")
    param_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Cabeçalho
    ws['A1'] = "ESTIMATIVA DO PROJETO"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Parâmetros editáveis
    ws['A3'] = "PARÂMETROS (editáveis)"
    ws['A3'].font = header_font
    
    ws['A4'] = "Valor/hora (R$):"
    ws['B4'] = valor_hora
    ws['B4'].font = param_font
    ws['B4'].fill = param_fill
    
    ws['A5'] = "Fator de segurança:"
    ws['B5'] = fator
    ws['B5'].font = param_font
    ws['B5'].fill = param_fill
    
    ws['A6'] = "Imposto (%):"
    ws['B6'] = imposto_pct * 100
    ws['B6'].font = param_font
    ws['B6'].fill = param_fill
    
    # Tarefas editáveis
    ws['A8'] = "TAREFAS (editáveis)"
    ws['A8'].font = header_font
    
    ws['A9'] = "Tarefa"
    ws['B9'] = "Horas"
    ws['A9'].font = header_font
    ws['B9'].font = header_font
    
    # Inserir tarefas
    linha_tarefa = 10
    for i, (tarefa, horas) in enumerate(tarefas):
        ws[f'A{linha_tarefa + i}'] = tarefa
        ws[f'B{linha_tarefa + i}'] = horas
    
    # Adicionar linhas extras para futuras tarefas
    for i in range(len(tarefas), len(tarefas) + 10):
        ws[f'A{linha_tarefa + i}'] = ""
        ws[f'B{linha_tarefa + i}'] = 0
    
    # Linha dos totais (com fórmulas)
    linha_total = linha_tarefa + len(tarefas) + 12
    
    ws[f'A{linha_total}'] = "TOTAIS (automáticos)"
    ws[f'A{linha_total}'].font = header_font
    
    ws[f'A{linha_total + 1}'] = "Total estimado (h):"
    ws[f'B{linha_total + 1}'] = f"=SUM(B{linha_tarefa}:B{linha_tarefa + len(tarefas) + 9})"
    ws[f'B{linha_total + 1}'].font = total_font
    ws[f'B{linha_total + 1}'].fill = total_fill
    
    ws[f'A{linha_total + 2}'] = "Com fator segurança (h):"
    ws[f'B{linha_total + 2}'] = f"=B{linha_total + 1}*B5"
    ws[f'B{linha_total + 2}'].font = total_font
    ws[f'B{linha_total + 2}'].fill = total_fill
    
    ws[f'A{linha_total + 3}'] = "Valor bruto (R$):"
    ws[f'B{linha_total + 3}'] = f"=B{linha_total + 2}*B4"
    ws[f'B{linha_total + 3}'].font = total_font
    ws[f'B{linha_total + 3}'].fill = total_fill
    ws[f'B{linha_total + 3}'].number_format = 'R$ #,##0.00'
    
    ws[f'A{linha_total + 4}'] = "Impostos (R$):"
    ws[f'B{linha_total + 4}'] = f"=B{linha_total + 3}*(B6/100)"
    ws[f'B{linha_total + 4}'].font = total_font
    ws[f'B{linha_total + 4}'].fill = total_fill
    ws[f'B{linha_total + 4}'].number_format = 'R$ #,##0.00'
    
    ws[f'A{linha_total + 5}'] = "Valor líquido (R$):"
    ws[f'B{linha_total + 5}'] = f"=B{linha_total + 3}-B{linha_total + 4}"
    ws[f'B{linha_total + 5}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'B{linha_total + 5}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws[f'B{linha_total + 5}'].number_format = 'R$ #,##0.00'
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    caminho_excel = os.path.join(pasta_projeto, "estimativa_projeto.xlsx")
    wb.save(caminho_excel)
    print(f"📂 Planilha dinâmica salva em '{caminho_excel}'")
    print("💡 Agora você pode editar tarefas, horas e parâmetros diretamente no Excel!")

def salvar_pdf(tarefas, horas_totais, horas_ajustadas, valor_bruto, valor_imposto, valor_liquido, pasta_projeto):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, "Relatório de Estimativa do Projeto", ln=True, align="C")

    pdf.ln(10)
    for t, h in tarefas:
        pdf.cell(0, 10, f"- {t}: {h}h", ln=True)

    pdf.ln(5)
    pdf.cell(0, 10, f"Total estimado: {horas_totais:.2f}h", ln=True)
    pdf.cell(0, 10, f"Com fator segurança: {horas_ajustadas:.2f}h", ln=True)
    pdf.cell(0, 10, f"Valor bruto: R$ {valor_bruto:.2f}", ln=True)
    pdf.cell(0, 10, f"Impostos: R$ {valor_imposto:.2f}", ln=True)
    pdf.cell(0, 10, f"Valor líquido: R$ {valor_liquido:.2f}", ln=True)

    caminho_pdf = os.path.join(pasta_projeto, "estimativa_projeto.pdf")
    pdf.output(caminho_pdf)
    print(f"📄 Relatório PDF salvo em '{caminho_pdf}'")

def calculadora_freela_projeto():
    print("=== Calculadora de Projeto (tarefas personalizadas) ===")

    # Nome do projeto e criação da pasta
    nome_projeto = input("Qual o nome do projeto? ").strip()
    pasta_projeto = nome_projeto.replace(" ", "_")  # Remove espaços para criar pasta
    
    if not os.path.exists(pasta_projeto):
        os.makedirs(pasta_projeto)
        print(f"📁 Pasta '{pasta_projeto}' criada!")
    else:
        print(f"📁 Usando pasta existente '{pasta_projeto}'")

    # Valor/hora
    valor_hora = float(input("Qual o valor/hora que você cobra? R$ ").replace(",", "."))

    # Inserção de tarefas
    tarefas = []
    print("\nDigite as tarefas e o tempo estimado de cada uma (em horas).")
    print("Quando terminar, aperte ENTER sem nada.\n")

    while True:
        tarefa = input("Tarefa (ou ENTER para finalizar): ").strip()
        if tarefa == "":
            break
        horas = float(input(f"Quantas horas para '{tarefa}'? ").replace(",", "."))
        tarefas.append((tarefa, horas))

    if not tarefas:
        print("⚠️ Nenhuma tarefa inserida. Saindo...")
        return

    # Fator de segurança
    fator_str = input("\nFator de segurança (padrão 1.3 = +30%): ").strip()
    fator = float(fator_str.replace(",", ".") if fator_str else 1.3)

    # Imposto
    imposto_str = input("Porcentagem de imposto (padrão 6%): ").strip()
    imposto = float(imposto_str.replace(",", ".") if imposto_str else 6.0) / 100

    # Cálculos
    horas_totais = sum(h for _, h in tarefas)
    horas_ajustadas = horas_totais * fator
    valor_bruto = horas_ajustadas * valor_hora
    valor_liquido = valor_bruto * (1 - imposto)
    valor_imposto = valor_bruto - valor_liquido

    # Relatório no terminal
    print("\n--- RELATÓRIO ---")
    for t, h in tarefas:
        print(f"- {t}: {h}h")
    print(f"\nTotal estimado: {horas_totais:.2f}h")
    print(f"Com fator de segurança ({fator:.2f}): {horas_ajustadas:.2f}h")
    print(f"\nValor total BRUTO do projeto: R$ {valor_bruto:.2f}")
    print(f"Impostos: R$ {valor_imposto:.2f}")
    print(f"Valor total LÍQUIDO (após impostos): R$ {valor_liquido:.2f}")

    # Salvando automaticamente os relatórios
    print("\n💾 Salvando relatórios...")
    salvar_excel(tarefas, horas_totais, horas_ajustadas, valor_bruto, valor_imposto, valor_liquido, pasta_projeto, valor_hora, fator, imposto)
    salvar_pdf(tarefas, horas_totais, horas_ajustadas, valor_bruto, valor_imposto, valor_liquido, pasta_projeto)
    print("✅ Relatórios salvos com sucesso!")


if __name__ == "__main__":
    calculadora_freela_projeto()
